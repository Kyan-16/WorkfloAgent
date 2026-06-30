"""
内容生成工具集

提供三个内容创作类工具，均继承自 tools.base.Tool：
- TextGeneratorTool: 文生文（调用 LLM 生成文本）
- ImageGeneratorTool: 文生图（模拟 / 真实 API）
- ReportGeneratorTool: 工单统计报表（PDF / Excel）
"""
import json
import logging
import os
import subprocess
import sys
import tempfile
import uuid
from datetime import datetime, timezone
from typing import Optional

from tools.base import Tool, ToolResult
from llm.base import ChatMessage

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# 工具1：文生文
# ---------------------------------------------------------------------------

class TextGeneratorTool(Tool):
    """根据用户需求生成各类文本内容，如文章、报告、邮件、文案等"""

    name = "generate_text"
    description = "根据用户需求生成各类文本内容，如文章、报告、邮件、文案等"
    parameters = {
        "type": "object",
        "properties": {
            "topic": {
                "type": "string",
                "description": "文本主题或标题",
            },
            "style": {
                "type": "string",
                "description": "写作风格（正式/幽默/简洁/详细），默认正式",
                "enum": ["正式", "幽默", "简洁", "详细"],
            },
            "length": {
                "type": "string",
                "description": "篇幅要求（短/中/长），默认中",
                "enum": ["短", "中", "长"],
            },
            "format": {
                "type": "string",
                "description": "输出格式（纯文本/Markdown），默认 Markdown",
                "enum": ["纯文本", "Markdown"],
            },
        },
        "required": ["topic"],
    }

    def __init__(self, llm=None):
        super().__init__()
        self._llm = llm

    @property
    def llm(self):
        return self._llm

    @llm.setter
    def llm(self, value):
        self._llm = value

    async def execute(
        self,
        topic: str,
        style: str = "正式",
        length: str = "中",
        format: str = "Markdown",
    ) -> ToolResult:
        if not self._llm:
            return ToolResult(
                success=False,
                error="TextGeneratorTool 未注入 LLM 实例，无法生成文本",
            )

        # 字数映射
        length_map = {"短": "约 300 字", "中": "约 800 字", "长": "约 1500 字"}

        system_prompt = (
            "你是一个专业的文本创作助手。请严格按照用户要求的风格、篇幅和格式输出。"
            "不要添加额外的说明或评论，直接输出正文。"
        )

        user_prompt = (
            f"请以【{style}】的风格，写一篇【{topic}】。\n"
            f"篇幅要求：{length_map.get(length, '约 800 字')}\n"
            f"输出格式：{format}\n\n"
            "请直接输出正文内容，不要包含额外的解释。"
        )

        try:
            messages = [
                ChatMessage(role="system", content=system_prompt),
                ChatMessage(role="user", content=user_prompt),
            ]
            response = await self._llm.generate(
                messages,
                temperature=0.8 if style == "幽默" else 0.6,
                max_tokens=2048,
            )

            if response.is_error:
                return ToolResult(
                    success=False,
                    error=f"LLM 返回错误: {response.content}",
                )

            return ToolResult(
                success=True,
                output={
                    "content": response.content,
                    "topic": topic,
                    "style": style,
                    "length": length,
                    "format": format,
                    "model": response.model,
                },
            )
        except Exception as e:
            logger.exception("文本生成失败")
            return ToolResult(success=False, error=f"文本生成失败: {str(e)}")


# ---------------------------------------------------------------------------
# 工具2：文生图
# ---------------------------------------------------------------------------

class ImageGeneratorTool(Tool):
    """根据文本描述生成图片"""

    name = "generate_image"
    description = "根据文本描述生成图片"
    parameters = {
        "type": "object",
        "properties": {
            "prompt": {
                "type": "string",
                "description": "图片描述文本",
            },
            "style": {
                "type": "string",
                "description": "风格（写实/卡通/插画/水墨），默认写实",
                "enum": ["写实", "卡通", "插画", "水墨"],
            },
            "size": {
                "type": "string",
                "description": "图片尺寸（如 1024x1024），默认 1024x1024",
                "pattern": "^\\d+x\\d+$",
            },
        },
        "required": ["prompt"],
    }

    def __init__(self):
        super().__init__()
        self._api_key = os.environ.get("IMAGE_API_KEY", "")

    async def execute(
        self,
        prompt: str,
        style: str = "写实",
        size: str = "1024x1024",
    ) -> ToolResult:
        enhanced_prompt = f"[{style}风格] {prompt}"

        if self._api_key:
            try:
                return await self._call_real_api(enhanced_prompt, style, size)
            except Exception as e:
                logger.warning(f"真实文生图 API 调用失败，回退模拟模式: {e}")

        # 模拟模式
        return self._mock_result(enhanced_prompt, style, size)

    async def _call_real_api(self, prompt: str, style: str, size: str) -> ToolResult:
        """
        调用外部文生图 API。
        当前作为扩展点预留，可对接 DALL·E / Stable Diffusion / 通义万相等。
        """
        import httpx

        api_url = os.environ.get("IMAGE_API_URL", "https://api.example.com/v1/images/generations")
        headers = {
            "Authorization": f"Bearer {self._api_key}",
            "Content-Type": "application/json",
        }
        payload = {
            "prompt": prompt,
            "size": size,
            "n": 1,
        }

        async with httpx.AsyncClient(timeout=60) as client:
            resp = await client.post(api_url, json=payload, headers=headers)
            resp.raise_for_status()
            data = resp.json()

        image_url = data.get("data", [{}])[0].get("url", "")
        return ToolResult(
            success=True,
            output={
                "image_url": image_url,
                "prompt": prompt,
                "style": style,
                "size": size,
                "mode": "real_api",
            },
        )

    def _mock_result(self, prompt: str, style: str, size: str) -> ToolResult:
        """返回模拟的图片 URL"""
        mock_id = uuid.uuid4().hex[:8]
        mock_url = f"https://image.mock.service/{mock_id}?prompt={prompt[:30]}&style={style}&size={size}"

        return ToolResult(
            success=True,
            output={
                "image_url": mock_url,
                "prompt": prompt,
                "style": style,
                "size": size,
                "mode": "simulated",
                "notice": "当前为模拟模式，返回的是模拟图片链接。"
                          "如需真实文生图能力，请设置环境变量 IMAGE_API_KEY。",
            },
        )


# ---------------------------------------------------------------------------
# 工具3：生成 PDF / 数据表
# ---------------------------------------------------------------------------

class ReportGeneratorTool(Tool):
    """生成工单统计报告或数据表，输出为 PDF 或 Excel 格式"""

    name = "generate_report"
    description = "生成工单统计报告或数据表，输出为 PDF 或 Excel 格式"
    parameters = {
        "type": "object",
        "properties": {
            "report_type": {
                "type": "string",
                "description": "报告类型（pdf/excel），默认 pdf",
                "enum": ["pdf", "excel"],
            },
            "title": {
                "type": "string",
                "description": "报告标题",
            },
            "data_source": {
                "type": "string",
                "description": "数据范围（全部工单/本周/本月），默认全部",
                "enum": ["全部工单", "本周", "本月"],
            },
            "include_chart": {
                "type": "boolean",
                "description": "是否包含图表（仅 PDF 支持），默认 false",
            },
        },
        "required": ["title"],
    }

    async def execute(
        self,
        title: str,
        report_type: str = "pdf",
        data_source: str = "全部工单",
        include_chart: bool = False,
    ) -> ToolResult:
        # 1. 获取工单数据
        tickets = await self._fetch_tickets(data_source)
        if tickets is None:
            return ToolResult(
                success=False,
                error="无法获取工单数据，请确认数据库连接正常。",
            )

        # 2. 确保输出目录存在
        reports_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            "data", "reports",
        )
        os.makedirs(reports_dir, exist_ok=True)

        # 3. 生成统计信息
        stats = self._compute_stats(tickets)

        # 4. 按类型生成
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_title = title.replace(" ", "_").replace("/", "_")

        if report_type == "pdf":
            return await self._generate_pdf(
                title, safe_title, timestamp, tickets, stats, reports_dir, include_chart,
            )
        else:
            return await self._generate_excel(
                title, safe_title, timestamp, tickets, stats, reports_dir,
            )

    # ---- 数据获取 ----

    async def _fetch_tickets(self, data_source: str):
        """从数据库查询工单"""
        try:
            from ticket_agent.repository import get_ticket_repository

            repo = get_ticket_repository()
            all_tickets = repo.list_all(limit=1000, offset=0)
            if not all_tickets:
                return []

            # 按时间范围过滤
            if data_source == "全部工单":
                return all_tickets

            now = datetime.now(timezone.utc)
            if data_source == "本月":
                filtered = [
                    t for t in all_tickets
                    if t.created_at and self._parse_datetime(t.created_at).month == now.month
                    and self._parse_datetime(t.created_at).year == now.year
                ]
                return filtered

            if data_source == "本周":
                # 本周一
                monday = now.date()
                monday = monday.replace(day=monday.day - monday.weekday())
                filtered = [
                    t for t in all_tickets
                    if t.created_at
                    and self._parse_datetime(t.created_at).date() >= monday
                ]
                return filtered

            return all_tickets
        except ImportError:
            logger.warning("无法导入 ticket_agent.repository，返回空数据")
            return []
        except Exception as e:
            logger.error(f"获取工单数据失败: {e}")
            return []

    @staticmethod
    def _parse_datetime(dt_str: str):
        """安全解析 ISO 时间字符串"""
        try:
            return datetime.fromisoformat(dt_str)
        except (ValueError, TypeError):
            return datetime.min.replace(tzinfo=timezone.utc)

    # ---- 统计 ----

    def _compute_stats(self, tickets) -> dict:
        """计算工单统计信息"""
        total = len(tickets)
        category_stats = {}
        status_stats = {}
        user_tickets = {}

        for t in tickets:
            cat = t.category.value if t.category else "未分类"
            category_stats[cat] = category_stats.get(cat, 0) + 1

            st = t.status.value if t.status else "未知"
            status_stats[st] = status_stats.get(st, 0) + 1

            uid = t.user_id or "anonymous"
            user_tickets[uid] = user_tickets.get(uid, 0) + 1

        # 找出提交最多工单的用户
        top_user = max(user_tickets, key=user_tickets.get) if user_tickets else "N/A"
        top_user_count = user_tickets.get(top_user, 0)

        return {
            "total": total,
            "category_stats": category_stats,
            "status_stats": status_stats,
            "top_user": top_user,
            "top_user_count": top_user_count,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    # ---- PDF ----

    @staticmethod
    def _find_unicode_font():
        """
        查找系统中可用的 Unicode / CJK 字体。
        返回 (font_path, font_name) 或 (None, None)。
        """
        # 常见的中文字体路径
        candidates = []
        if os.name == "nt":  # Windows
            win_fonts = os.environ.get("WINDIR", r"C:\Windows")
            candidates = [
                (os.path.join(win_fonts, "Fonts", "simsun.ttc"), "SimSun"),
                (os.path.join(win_fonts, "Fonts", "msyh.ttc"), "MicrosoftYaHei"),
                (os.path.join(win_fonts, "Fonts", "simhei.ttf"), "SimHei"),
            ]
        else:  # Linux / macOS
            candidates = [
                ("/usr/share/fonts/truetype/wqy/wqy-microhei.ttc", "WenQuanYiMicroHei"),
                ("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc", "NotoSansCJK"),
                ("/System/Library/Fonts/PingFang.ttc", "PingFang"),
                ("/System/Library/Fonts/STHeiti Light.ttc", "STHeiti"),
            ]

        for path, name in candidates:
            if os.path.exists(path):
                return path, name
        return None, None

    async def _generate_pdf(
        self,
        title: str,
        safe_title: str,
        timestamp: str,
        tickets,
        stats: dict,
        reports_dir: str,
        include_chart: bool,
    ) -> ToolResult:
        try:
            from fpdf import FPDF
        except ImportError:
            self._ensure_dependency("fpdf2")

        from fpdf import FPDF
        from fpdf.errors import FPDFUnicodeEncodingException

        pdf = FPDF()
        pdf.add_page()

        # 检测并加载 Unicode 字体（支持中文）
        font_path, font_name = self._find_unicode_font()
        if font_path:
            # fpdf2 需要为每个变体分别注册；这里用同一字体文件注册所有变体
            for style in ("", "B", "I", "BI"):
                try:
                    pdf.add_font(font_name, style, font_path)
                except Exception:
                    pass  # 部分变体可能不受支持，忽略
            title_font = font_name
            body_font = font_name
        else:
            # 无 Unicode 字体时的回退：使用 Helvetica（仅支持 Latin-1）
            title_font = "Helvetica"
            body_font = "Helvetica"
            logger.warning("未找到 Unicode 字体，PDF 中的中文可能显示为乱码")

        # 渲染内容（含编码容错：无 Unicode 字体时自动降级为纯英文 PDF）
        try:
            result = self._render_pdf_content(
                pdf, title, title_font, body_font, tickets, stats,
            )
            if result:
                return result
        except FPDFUnicodeEncodingException:
            logger.warning("PDF 编码失败（当前字体不支持中文），回退到 ASCII 版本")

        # 回退：纯英文 PDF（任何字体都可渲染）
        pdf2 = FPDF()
        pdf2.add_page()
        pdf2.set_font("Helvetica", "B", 16)
        pdf2.cell(0, 15, title, align="C", new_x="LMARGIN", new_y="NEXT")
        pdf2.ln(5)
        pdf2.set_font("Helvetica", "", 10)
        pdf2.cell(0, 8, f"Generated: {stats['generated_at']}", new_x="LMARGIN", new_y="NEXT")
        pdf2.cell(0, 8, f"Total tickets: {len(tickets)}", new_x="LMARGIN", new_y="NEXT")
        pdf2.ln(5)
        pdf2.set_font("Helvetica", "B", 12)
        pdf2.cell(0, 10, "Category Statistics", new_x="LMARGIN", new_y="NEXT")
        pdf2.set_font("Helvetica", "", 10)
        for cat, count in sorted(stats["category_stats"].items(), key=lambda x: -x[1]):
            pdf2.cell(0, 8, f"  {cat}: {count}", new_x="LMARGIN", new_y="NEXT")

        filename = f"{safe_title}_{timestamp}.pdf"
        filepath = os.path.join(reports_dir, filename)
        pdf2.output(filepath)

        return ToolResult(
            success=True,
            output={
                "file_path": os.path.abspath(filepath),
                "file_name": filename,
                "report_type": "pdf",
                "ticket_count": len(tickets),
                "generated_at": stats["generated_at"],
            },
        )

    # ---- PDF ----

    def _render_pdf_content(
        self, pdf, title: str, title_font: str, body_font: str,
        tickets, stats: dict,
    ) -> ToolResult | None:
        """渲染 PDF 中文内容（可能因缺字体抛出 FPDFUnicodeEncodingException）"""
        from fpdf import FPDF

        pdf.set_font(title_font, "B", 18)
        pdf.cell(0, 15, title, align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)
        pdf.set_font(body_font, "", 11)
        pdf.cell(0, 8, f"生成时间: {stats['generated_at']}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 8, f"数据范围: {len(tickets)} 条工单", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)
        pdf.set_font(body_font, "B", 13)
        pdf.cell(0, 10, "一、按分类统计", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font(body_font, "", 11)
        for cat, count in sorted(stats["category_stats"].items(), key=lambda x: -x[1]):
            pdf.cell(0, 8, f"  {cat}: {count} 个工单", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)
        pdf.set_font(body_font, "B", 13)
        pdf.cell(0, 10, "二、按状态统计", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font(body_font, "", 11)
        for st, count in sorted(stats["status_stats"].items(), key=lambda x: -x[1]):
            pdf.cell(0, 8, f"  {st}: {count} 个工单", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)
        pdf.set_font(body_font, "B", 13)
        pdf.cell(0, 10, "三、活跃用户", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font(body_font, "", 11)
        pdf.cell(0, 8, f"  提交工单最多的用户: {stats['top_user']} ({stats['top_user_count']} 个工单)", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)
        pdf.set_font(body_font, "B", 13)
        pdf.cell(0, 10, "四、工单明细", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font(body_font, "B", 8)
        col_w = [30, 22, 30, 20, 40]
        headers = ["工单ID", "用户", "分类", "状态", "内容摘要"]
        for i, h in enumerate(headers):
            pdf.cell(col_w[i], 7, h, border=1)
        pdf.ln()
        pdf.set_font(body_font, "", 8)
        for t in tickets[:50]:
            row = [t.ticket_id[-12:], t.user_id or "-",
                   t.category.value if t.category else "-",
                   t.status.value if t.status else "-",
                   t.content[:28] + "..." if len(t.content) > 28 else t.content]
            for i, val in enumerate(row):
                pdf.cell(col_w[i], 6, val, border=1)
            pdf.ln()
        if len(tickets) > 50:
            pdf.set_font(body_font, "I", 9)
            pdf.cell(0, 8, f"... 仅展示前 50 条，共 {len(tickets)} 条", new_x="LMARGIN", new_y="NEXT")
        return None

    # ---- Excel ----

    async def _generate_excel(
        self,
        title: str,
        safe_title: str,
        timestamp: str,
        tickets,
        stats: dict,
        reports_dir: str,
    ) -> ToolResult:
        try:
            import openpyxl
        except ImportError:
            self._ensure_dependency("openpyxl")
            import openpyxl

        from openpyxl.styles import Font, Alignment, PatternFill

        wb = openpyxl.Workbook()

        # ---------- Sheet1：工单明细 ----------
        ws1 = wb.active
        ws1.title = "工单明细"

        # 标题行
        ws1.merge_cells("A1:F1")
        ws1["A1"] = title
        ws1["A1"].font = Font(bold=True, size=14)
        ws1["A2"] = f"生成时间: {stats['generated_at']}   工单总数: {len(tickets)}"
        ws1["A2"].font = Font(italic=True, size=10)

        headers = ["工单ID", "用户ID", "分类", "状态", "创建时间", "内容"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, t in enumerate(tickets, 5):
            ws1.cell(row=i, column=1, value=t.ticket_id)
            ws1.cell(row=i, column=2, value=t.user_id or "")
            ws1.cell(row=i, column=3, value=t.category.value if t.category else "")
            ws1.cell(row=i, column=4, value=t.status.value if t.status else "")
            ws1.cell(row=i, column=5, value=t.created_at or "")
            ws1.cell(row=i, column=6, value=t.content)

        # 列宽调整
        ws1.column_dimensions["A"].width = 28
        ws1.column_dimensions["B"].width = 14
        ws1.column_dimensions["C"].width = 10
        ws1.column_dimensions["D"].width = 12
        ws1.column_dimensions["E"].width = 22
        ws1.column_dimensions["F"].width = 50

        # ---------- Sheet2：分类统计汇总 ----------
        ws2 = wb.create_sheet("分类统计")

        ws2.merge_cells("A1:C1")
        ws2["A1"] = f"{title} — 分类统计汇总"
        ws2["A1"].font = Font(bold=True, size=13)

        # 按分类
        ws2["A3"] = "分类"
        ws2["B3"] = "工单数"
        ws2["C3"] = "占比"
        for h in ("A3", "B3", "C3"):
            ws2[h].font = header_font
            ws2[h].fill = header_fill
            ws2[h].alignment = Alignment(horizontal="center")

        total = stats["total"] or 1
        for i, (cat, cnt) in enumerate(
            sorted(stats["category_stats"].items(), key=lambda x: -x[1]), 4
        ):
            ws2.cell(row=i, column=1, value=cat)
            ws2.cell(row=i, column=2, value=cnt)
            ws2.cell(row=i, column=3, value=f"{cnt / total * 100:.1f}%")

        # 按状态
        status_start_row = len(stats["category_stats"]) + 6
        ws2.cell(row=status_start_row, column=1, value="状态").font = header_font
        ws2.cell(row=status_start_row, column=2, value="工单数").font = header_font
        ws2.cell(row=status_start_row, column=3, value="占比").font = header_font
        for h in ("A", "B", "C"):
            cell = ws2[f"{h}{status_start_row}"]
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, (st, cnt) in enumerate(
            sorted(stats["status_stats"].items(), key=lambda x: -x[1]),
            status_start_row + 1,
        ):
            ws2.cell(row=i, column=1, value=st)
            ws2.cell(row=i, column=2, value=cnt)
            ws2.cell(row=i, column=3, value=f"{cnt / total * 100:.1f}%")

        ws2.column_dimensions["A"].width = 14
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 10

        # 保存
        filename = f"{safe_title}_{timestamp}.xlsx"
        filepath = os.path.join(reports_dir, filename)
        wb.save(filepath)

        return ToolResult(
            success=True,
            output={
                "file_path": os.path.abspath(filepath),
                "file_name": filename,
                "report_type": "excel",
                "sheets": ["工单明细", "分类统计"],
                "ticket_count": len(tickets),
                "generated_at": stats["generated_at"],
            },
        )

    # ---- 依赖安装 ----

    @staticmethod
    def _ensure_dependency(package: str):
        """按需安装 Python 包"""
        logger.info(f"正在安装依赖: {package}")
        result = subprocess.run(
            [sys.executable, "-m", "pip", "install", package],
            capture_output=True, text=True, timeout=120,
        )
        if result.returncode != 0:
            raise RuntimeError(
                f"安装 {package} 失败: {result.stderr[:200]}"
            )
        logger.info(f"依赖安装完成: {package}")


# ---------------------------------------------------------------------------
# 注册函数
# ---------------------------------------------------------------------------

def register_content_tools(tool_registry, llm=None):
    """
    注册内容生成工具到指定的 ToolRegistry。

    :param tool_registry: ToolRegistry 实例
    :param llm: (可选) LLMBase 实例，TextGeneratorTool 需要
    """
    text_tool = TextGeneratorTool(llm=llm)
    tool_registry.register(text_tool)
    tool_registry.register(ImageGeneratorTool())
    tool_registry.register(ReportGeneratorTool())
